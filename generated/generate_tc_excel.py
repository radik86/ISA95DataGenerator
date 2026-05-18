"""
Generate ISA-95 Functional Relationship Test Cases Excel workbook.
Run with:  py generate_tc_excel.py
Output:    generated/ISA95_Test_Cases.xlsx

Column layout (8 columns):
  A  TC ID
  B  Functional Description
  C  Relationship Chain
  D  Anchor Entity Type
  E  Test Parameters          ← Anchor Entity ID  +  Timestamp / Time Window
  F  Expected Results         ← Entity Counts  +  Key IDs  +  Latest Values + UoM
  G  Validation SQL
  H  Validation Criteria
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
ODD_FILL    = PatternFill("solid", fgColor="DEEAF1")
EVEN_FILL   = PatternFill("solid", fgColor="FFFFFF")
SQL_FILL    = PatternFill("solid", fgColor="F2F2F2")
PARAM_FILL_ODD  = PatternFill("solid", fgColor="E2EFDA")   # light green – test params
PARAM_FILL_EVEN = PatternFill("solid", fgColor="F0F7EC")
EXP_FILL_ODD    = PatternFill("solid", fgColor="FCE4D6")   # light orange – expected results
EXP_FILL_EVEN   = PatternFill("solid", fgColor="FEF4EF")

WHITE  = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
NORMAL = Font(name="Calibri", size=10)
BOLD   = Font(name="Calibri", size=10, bold=True)
TCID   = Font(name="Calibri", size=10, bold=True, color="1F4E79")
MONO   = Font(name="Courier New", size=9)

THIN   = Side(style="thin", color="B8CCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP   = Alignment(wrap_text=True, vertical="top")
MID    = Alignment(wrap_text=True, vertical="center", horizontal="center")

# ---------------------------------------------------------------------------
# Column definitions  (header, width)
# ---------------------------------------------------------------------------
COLUMNS = [
    ("TC ID",                    12),
    ("Functional Description",   52),
    ("Relationship Chain",       44),
    ("Anchor Entity Type",       24),
    ("Test Parameters",          62),   # anchor IDs + timestamp / time window
    ("Expected Results",         88),   # counts + key IDs + latest values + UoM
    ("Validation SQL",           90),
    ("Validation Criteria",      60),
]

# Column indices (1-based)
COL_TC        = 1
COL_DESC      = 2
COL_CHAIN     = 3
COL_TYPE      = 4
COL_PARAMS    = 5   # green
COL_EXPECTED  = 6   # orange
COL_SQL       = 7
COL_CRITERIA  = 8

# ---------------------------------------------------------------------------
# Test case records
# Each tuple:
#   (tc_id, func_desc, rel_chain, anchor_type,
#    test_params,      expected_results,
#    validation_sql,   validation_criteria)
#
# test_params  = anchor entity ID(s)  +  relevant timestamps / time windows
# expected_results = entity counts  +  key IDs / names  +  latest values + UoM
# ---------------------------------------------------------------------------

TC = [
# ── TC-001 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-001",
"Verify that a Production operations event class fans out to all its event definitions and each definition has a linked event record. Validates the master-data backbone for production scheduling events.",
"Operations Event Class → Operation Event Definition → Operations Event Record",
"Operations Event Class",
# Test Parameters
"Event Class ID: Production",
# Expected Results
"Counts:\n  Operation Event Definitions: 6\n  Operations Event Records: 6\n\n"
"Definition IDs:\n  OED-PROD-BAKE-END, OED-PROD-BAKE-START\n  OED-PROD-MIX-END, OED-PROD-MIX-START\n  OED-PROD-PACK-END, OED-PROD-PACK-START\n\n"
"Record IDs: OER-T-008 … OER-T-013",
"""SELECT c.OperationsEventClassID, d.Id AS DefinitionId, r.Id AS RecordId
FROM OperationsEventClasses c
LEFT JOIN OperationEventDefinitions d ON d.EventCategory = c.OperationsEventClassID
LEFT JOIN OperationsEventRecords r ON r.OperationsEventDefinitionID = d.Id
WHERE c.OperationsEventClassID = 'Production'
ORDER BY d.Id, r.Id;""",
"All 6 OED records reference class 'Production'.\nAll 6 OER records reference one of those OEDs.\nAll records have non-empty sourceTimeStamp.",
),
# ── TC-002 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-002",
"Verify that an Operations Request decomposes into Segment Requirements carrying both material and equipment constraints. Validates request-level planning completeness.",
"Operations Request → Segment Requirement → Material Requirement; Segment Requirement → Equipment Requirement",
"Operations Request",
# Test Parameters
"Operations Request ID:\n  OR-PLANT01MUNICH-LINE-01-202601270827-801\n\n"
"Snapshot timestamp: 2026-01-01T06:30:00.000Z",
# Expected Results
"Counts:\n  Segment Requirements: 3\n  Material Requirements: 7\n  Equipment Requirements: 3\n\n"
"Material names: Flour, Water, Yeast, Salt, Dough, Paper, BREAD-750G\n\n"
"Equipment IDs: EQ-MIX-01, EQ-OVEN-01, EQ-PACKAGE-01",
"""DECLARE @orid NVARCHAR(200) = 'OR-PLANT01MUNICH-LINE-01-202601270827-801';
SELECT RecordId AS SegReqId
FROM GenericDataStores
WHERE StoreName='segmentRequirements'
  AND JSON_VALUE(DataJson,'$.operationsRequestId')=@orid;
SELECT RecordId AS MatReqId
FROM GenericDataStores
WHERE StoreName='segmentMaterialRequirements'
  AND JSON_VALUE(DataJson,'$.segmentRequirementId') IN (
    SELECT RecordId FROM GenericDataStores
    WHERE StoreName='segmentRequirements'
      AND JSON_VALUE(DataJson,'$.operationsRequestId')=@orid);
SELECT RecordId AS EqReqId
FROM GenericDataStores
WHERE StoreName='segmentEquipmentRequirements'
  AND JSON_VALUE(DataJson,'$.segmentRequirementId') IN (
    SELECT RecordId FROM GenericDataStores
    WHERE StoreName='segmentRequirements'
      AND JSON_VALUE(DataJson,'$.operationsRequestId')=@orid);""",
"Exactly 3 SegReqs reference the OR.\nExactly 7 MatReqs across those SegReqs.\nExactly 3 EqReqs across those SegReqs.\nAll records have non-empty sourceTimeStamp.",
),
# ── TC-003 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-003",
"Verify that an Operations Response links back to its originating Operations Request and that all Segment Responses are accessible. Validates planning–execution traceability.",
"Operations Response → Operations Request; Operations Response → Segment Response",
"Operations Response",
# Test Parameters
"Operations Response ID:\n  OPS-RESP-PLANT01MUNICH-LINE-01-202604161209-216\n\n"
"Snapshot timestamp: 2026-01-01T06:30:00.000Z",
# Expected Results
"Counts:\n  Operations Requests linked: 1\n  Segment Responses: 6\n\n"
"Operations Request ID:\n  OR-PLANT01MUNICH-LINE-01-202601270827-801",
"""SELECT JSON_VALUE(DataJson,'$.operationsRequestId') AS OpsReqId
FROM GenericDataStores
WHERE StoreName='operationsResponses'
  AND RecordId='OPS-RESP-PLANT01MUNICH-LINE-01-202604161209-216';
SELECT RecordId AS SegRespId
FROM GenericDataStores
WHERE StoreName='segmentResponses'
  AND JSON_VALUE(DataJson,'$.operationsResponseId')=
      'OPS-RESP-PLANT01MUNICH-LINE-01-202604161209-216';""",
"Exactly 1 Operations Request linked.\nExactly 6 Segment Responses linked.\nAll records have non-empty sourceTimeStamp.",
),
# ── TC-004 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-004",
"Verify that a Material Lot is associated with its sublots and lot-level quality properties. Validates quality and traceability data completeness for a finished-goods lot.",
"Material Lot → Material Sublot; Material Lot → Material Lot Property",
"Material Lot",
# Test Parameters
"Material Lot ID:\n  LOT-PLANT01MUNICH-LINE-01-202601011230-BREAD-750G-R1\n\n"
"Lot production timestamp: 2026-01-01T11:30:00.000Z",
# Expected Results
"Counts:\n  Material Sublots: 4\n  Material Lot Properties: 1",
"""SELECT RecordId AS SublotId
FROM GenericDataStores
WHERE StoreName='materialSublots'
  AND JSON_VALUE(DataJson,'$.materialLotId')=
      'LOT-PLANT01MUNICH-LINE-01-202601011230-BREAD-750G-R1';
SELECT RecordId AS LotPropertyId
FROM GenericDataStores
WHERE StoreName='materialLotProperties'
  AND JSON_VALUE(DataJson,'$.materialLotId')=
      'LOT-PLANT01MUNICH-LINE-01-202601011230-BREAD-750G-R1';""",
"Exactly 4 sublots linked.\nExactly 1 lot property linked.\nAll records have non-empty sourceTimeStamp.",
),
# ── TC-005 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-005",
"Verify that a Test Result traces to its Material Lot and that the lot's sublots are accessible from the test result. Validates quality-to-traceability linkage.",
"Test Result → Material Lot → Material Sublot",
"Test Result",
# Test Parameters
"Test Result ID:\n  TEST-LOT-PLANT01MUNICH-LINE-01-202601011230-BREAD-750G-R1-083\n\n"
"Test Result timestamp: 2026-01-01T10:50:00.000Z\n"
"Material Lot timestamp: 2026-01-01T11:30:00.000Z",
# Expected Results
"Counts:\n  Material Lots linked: 1\n  Material Sublots on lot: 4\n\n"
"Material Lot ID:\n  LOT-PLANT01MUNICH-LINE-01-202601011230-BREAD-750G-R1",
"""SELECT JSON_VALUE(DataJson,'$.materialLotId') AS LotId
FROM GenericDataStores
WHERE StoreName='testResults'
  AND RecordId='TEST-LOT-PLANT01MUNICH-LINE-01-202601011230-BREAD-750G-R1-083';""",
"Exactly 1 Material Lot linked.\nExactly 4 sublots on that lot.\nAll records have non-empty sourceTimeStamp.",
),
# ── TC-006 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-006",
"Verify that an Operations Response resolves all consumed material actuals across its segment responses. Validates material consumption reporting for a full production response.",
"Operations Response → Segment Response → Material Actual",
"Operations Response",
# Test Parameters
"Operations Response ID:\n  OPS-RESP-PLANT01MUNICH-LINE-01-202604161209-216\n\n"
"Data time range: 2026-01-01T06:30:00Z – 2026-01-01T12:30:00Z",
# Expected Results
"Counts:\n  Segment Responses: 6\n  Material Actuals: 14\n\n"
"Segment Response IDs:\n"
"  SEG-RESP-…-202601010730-RUN1-658\n"
"  SEG-RESP-…-202601010930-RUN1-971\n"
"  SEG-RESP-…-202601010930-RUN2-903\n"
"  SEG-RESP-…-202601011130-RUN1-425\n"
"  SEG-RESP-…-202601011130-RUN2-027\n"
"  SEG-RESP-…-202601011330-RUN2-235",
"""SELECT ma.RecordId AS MatActualId,
       JSON_VALUE(ma.DataJson,'$.segmentResponseId') AS SegRespId
FROM GenericDataStores ma
WHERE ma.StoreName='segmentMaterialActuals'
  AND JSON_VALUE(ma.DataJson,'$.segmentResponseId') IN (
    SELECT RecordId FROM GenericDataStores
    WHERE StoreName='segmentResponses'
      AND JSON_VALUE(DataJson,'$.operationsResponseId')=
          'OPS-RESP-PLANT01MUNICH-LINE-01-202604161209-216');""",
"Exactly 6 SegResps linked.\nExactly 14 MatActuals across those SegResps.\nAll timestamps within 2026-01-01T06:30:00Z–2026-01-01T12:30:00Z.",
),
# ── TC-007 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-007",
"Verify that an Operations Response resolves the distinct equipment used in production via the equipment-actual chain. Validates equipment utilisation reporting from execution data.",
"Operations Response → Segment Response → Equipment Actual → Equipment",
"Operations Response",
# Test Parameters
"Operations Response ID:\n  OPS-RESP-PLANT01MUNICH-LINE-01-202604161209-216\n\n"
"Data time range: 2026-01-01T06:30:00Z – 2026-01-01T12:30:00Z",
# Expected Results
"Counts:\n  Segment Responses: 6\n  Equipment Actuals: 6\n  Equipment (distinct): 3\n\n"
"Equipment IDs: EQ-MIX-01, EQ-OVEN-01, EQ-PACKAGE-01\n\n"
"Sample EqActual: EQ-ACT-…-202601010730-EQ-MIX-01-36-Production-2-Hours (+ 5 more)",
"""SELECT ea.RecordId,
       JSON_VALUE(ea.DataJson,'$.equipmentId') AS EquipmentId
FROM GenericDataStores ea
WHERE ea.StoreName='segmentEquipmentActuals'
  AND JSON_VALUE(ea.DataJson,'$.segmentResponseId') IN (
    SELECT RecordId FROM GenericDataStores
    WHERE StoreName='segmentResponses'
      AND JSON_VALUE(DataJson,'$.operationsResponseId')=
          'OPS-RESP-PLANT01MUNICH-LINE-01-202604161209-216');""",
"Exactly 6 EqActuals.\n3 distinct Equipment IDs resolved.\nAll timestamps within expected window.",
),
# ── TC-008 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-008",
"Extend TC-002 by resolving each Equipment Requirement to its master-data Equipment record. Validates that planning constraints reference real, active equipment.",
"Operations Request → Segment Requirement → Equipment Requirement → Equipment",
"Operations Request",
# Test Parameters
"Operations Request ID:\n  OR-PLANT01MUNICH-LINE-01-202601270827-801\n\n"
"Snapshot timestamp: 2026-01-01T06:30:00.000Z",
# Expected Results
"Counts:\n  Segment Requirements: 3\n  Material Requirements: 7\n  Equipment Requirements: 3\n  Equipment (distinct): 3\n\n"
"Equipment IDs: EQ-MIX-01, EQ-OVEN-01, EQ-PACKAGE-01",
"""SELECT e.Id AS EquipmentId, er.RecordId AS EqReqId
FROM GenericDataStores er
JOIN Equipments e ON e.Id = JSON_VALUE(er.DataJson,'$.equipmentId')
WHERE er.StoreName='segmentEquipmentRequirements'
  AND JSON_VALUE(er.DataJson,'$.segmentRequirementId') IN (
    SELECT RecordId FROM GenericDataStores
    WHERE StoreName='segmentRequirements'
      AND JSON_VALUE(DataJson,'$.operationsRequestId')=
          'OR-PLANT01MUNICH-LINE-01-202601270827-801');""",
"Each EqReq maps to exactly 1 Equipment record.\nAll 3 equipment IDs exist in Equipments table.\nAll records have non-empty sourceTimeStamp.",
),
# ── TC-009 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-009",
"Verify that an alarm-type Operations Event traces through its record entries to the associated Segment Response. Validates the alarm-to-execution chain for process incidents.",
"Operations Event → Operations Event Record → Operations Event Record Entry → Segment Response",
"Operations Event",
# Test Parameters
"Operations Event ID:\n  OPS-EVENT-SEG-RESP-PLANT01MUNICH-LINE-01-202601012230-RUN1-841-OED-PROC-MIXERR-160\n\n"
"Alarm timestamp: 2026-01-01T21:37:15.000Z\n"
"Record Entry timestamp: 2026-01-01T21:42:15.000Z\n"
"Segment Response timestamp: 2026-01-01T21:30:00.000Z",
# Expected Results
"Counts:\n  Ops Event Definitions: 1\n  Ops Event Records: 1\n  Record Entries: 1\n  Segment Responses: 1\n\n"
"Definition ID: OED-PROC-MIXERR\n\n"
"Segment Response ID:\n  SEG-RESP-PLANT01MUNICH-LINE-01-202601012230-RUN1-841",
"""SELECT oe.RecordId,
       JSON_VALUE(oe.DataJson,'$.operationsEventDefinitionId') AS OEDefId
FROM GenericDataStores oe
WHERE oe.StoreName='operationsEvents'
  AND oe.RecordId='OPS-EVENT-SEG-RESP-PLANT01MUNICH-LINE-01-202601012230-RUN1-841-OED-PROC-MIXERR-160';
SELECT oee.RecordId,
       JSON_VALUE(oee.DataJson,'$.segmentResponseId') AS SegRespId
FROM GenericDataStores oee
WHERE oee.StoreName='operationsEventEntries'
  AND JSON_VALUE(oee.DataJson,'$.operationsEventId')=
      'OPS-EVENT-SEG-RESP-PLANT01MUNICH-LINE-01-202601012230-RUN1-841-OED-PROC-MIXERR-160';""",
"Chain resolves in exactly 1 record per hop.\noperationsEventType = 'Alarm'.\nSeg Response exists and matches anchor.",
),
# ── TC-010 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-010",
"Extend TC-009 to resolve the Equipment that was active during the alarmed Segment Response. Validates end-to-end alarm-to-equipment attribution.",
"Operations Event → Segment Response → Equipment Actual → Equipment",
"Operations Event",
# Test Parameters
"Operations Event ID:\n  OPS-EVENT-SEG-RESP-PLANT01MUNICH-LINE-01-202601012230-RUN1-841-OED-PROC-MIXERR-160\n\n"
"Alarm timestamp: 2026-01-01T21:37:15.000Z\n"
"Segment Response timestamp: 2026-01-01T21:30:00.000Z",
# Expected Results
"Counts:\n  Segment Responses: 1\n  Equipment Actuals: 1\n  Equipment: 1\n\n"
"Segment Response ID:\n  SEG-RESP-PLANT01MUNICH-LINE-01-202601012230-RUN1-841\n\n"
"Equipment ID: EQ-MIX-01",
"""SELECT ea.RecordId,
       JSON_VALUE(ea.DataJson,'$.equipmentId') AS EquipId
FROM GenericDataStores ea
WHERE ea.StoreName='segmentEquipmentActuals'
  AND JSON_VALUE(ea.DataJson,'$.segmentResponseId')=
      'SEG-RESP-PLANT01MUNICH-LINE-01-202601012230-RUN1-841';""",
"Chain terminates at EQ-MIX-01.\nAll records have non-empty sourceTimeStamp.",
),
# ── TC-011 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-011",
"Verify that every alarm-typed Operations Event in the dataset links to exactly one Operations Event Definition. Validates referential integrity of alarm data across the entire snapshot.",
"Operations Event (type=Alarm) → Operations Event Definition",
"Operations Event Type",
# Test Parameters
"Filter: operationsEventType = 'Alarm'\nFilter: category = 'Process'\nScope: full dataset snapshot\n\n"
"Sample event timestamp: 2026-01-01T21:37:15.000Z",
# Expected Results
"Counts:\n  Alarm events in snapshot: 1,898\n  Orphaned alarms (no definition): 0\n\n"
"Sample: OPS-EVENT-…-OED-PROC-MIXERR-160 → OED-PROC-MIXERR",
"""SELECT COUNT(*) AS AlarmCount
FROM GenericDataStores
WHERE StoreName='operationsEvents'
  AND JSON_VALUE(DataJson,'$.operationsEventType')='Alarm';
-- Orphan check (must return 0):
SELECT COUNT(*) AS Orphans
FROM GenericDataStores oe
WHERE oe.StoreName='operationsEvents'
  AND JSON_VALUE(oe.DataJson,'$.operationsEventType')='Alarm'
  AND NOT EXISTS (
    SELECT 1 FROM OperationEventDefinitions d
    WHERE d.Id = JSON_VALUE(oe.DataJson,'$.operationsEventDefinitionId'));""",
"Total alarm count = 1,898.\nOrphan count = 0 (every alarm has a valid definition).\nAll records have non-empty sourceTimeStamp.",
),
# ── TC-012 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-012",
"Verify that the latest sensor reading for a specific equipment property (Mixer Motor Current) is retrievable from historical tracking data. Validates the latest-value query pattern for equipment telemetry.",
"Equipment → Equipment Property (latest record)",
"Equipment",
# Test Parameters
"Equipment ID: EQ-MIX-01\n"
"Property ID: EP-MIX-CUR\n"
"Query type: latest record\n\n"
"Latest snapshot timestamp: 2026-05-05T09:51:00.000Z",
# Expected Results
"Counts:\n  Equipment Property records returned: 1\n\n"
"RecordId:\n  PROP-TRACK-PLANT01MUNICH-LINE-01-EQ-MIX-01-EP-MIX-CUR\n\n"
"Property description: MixerMotorCurrentA\n\n"
"Latest value: 93.84 A\n"
"Timestamp: 2026-05-05T09:51:00.000Z",
"""SELECT TOP 1 RecordId,
    CAST(DataJson AS NVARCHAR(MAX)) AS FullJson
FROM GenericDataStores
WHERE StoreName='equipmentPropertyTracking'
  AND RecordId LIKE
      'PROP-TRACK-EQ-ACT-PLANT01MUNICH-LINE-01-2026%EQ-MIX-01%EP-MIX-CUR%'
ORDER BY RecordId DESC;""",
"Only 1 record returned.\nRecord belongs to EQ-MIX-01 / EP-MIX-CUR.\nValue = 93.84, UoM = A.\ncreatedTimestamp = '2026-05-05 09:51:00'.",
),
# ── TC-013 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-013",
"Verify that all sensor readings for Mixer Motor Current within a defined 2-hour window are returned with no records outside the range. Validates time-range filtering for equipment telemetry queries.",
"Equipment → Equipment Property (time-range filter)",
"Equipment",
# Test Parameters
"Equipment ID: EQ-MIX-01\n"
"Property ID: EP-MIX-CUR\n"
"Sampling interval: 5 seconds\n\n"
"Time window (inclusive):\n"
"  From: 2026-01-01T06:30:00.000Z\n"
"  To:   2026-01-01T08:30:00.000Z",
# Expected Results
"Counts:\n  Equipment Property records in window: 1,442\n\n"
"RecordId prefix:\n  PROP-TRACK-PLANT01MUNICH-LINE-01-EQ-MIX-01-EP-MIX-CUR\n\n"
"Window span: 2 hours × 720 samples/hour (5s) = 1,440 expected\n"
"(+2 boundary records inclusive)",
"""SELECT COUNT(*) AS RecordCount
FROM GenericDataStores
WHERE StoreName='equipmentPropertyTracking'
  AND RecordId LIKE
      'PROP-TRACK-EQ-ACT-PLANT01MUNICH-LINE-01-20260101%EQ-MIX-01%EP-MIX-CUR%'
  AND RIGHT(RecordId,14) BETWEEN '20260101063000' AND '20260101083000';""",
"Record count = 1,442.\nAll records are for EQ-MIX-01 / EP-MIX-CUR.\nAll createdTimestamp values are within the window.\nNo out-of-range records returned.",
),
# ── TC-014 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-014",
"Verify that a MaintenanceExecution Operations Event traces to the correct Definition, through record-entry chain, and to the Equipment that underwent maintenance. Validates MNT event-to-equipment attribution.",
"Operations Event (category=MaintenanceExecution) → Definition; → Segment Response → Equipment Actual → Equipment",
"Operations Event",
# Test Parameters
"Operations Event ID:\n  OPS-EVENT-MNT-SRESP-MNT-RESP-20260416115813-1-OED-MNT-START-413\n\n"
"Event timestamp: 2026-02-09T15:39:26.000Z\n"
"Segment Response timestamp: 2026-02-09T14:30:00.000Z",
# Expected Results
"Counts:\n  Ops Event Definitions: 1\n  Ops Event Records: 1\n  Record Entries: 1\n  Segment Responses: 1\n  Equipment Actuals: 1\n  Equipment: 1\n\n"
"Definition ID: OED-MNT-START\n"
"Segment Response ID: MNT-SRESP-MNT-RESP-20260416115813-1\n"
"Equipment ID: EQ-MIX-01\n"
"Event category: MaintenanceExecution",
"""SELECT JSON_VALUE(DataJson,'$.operationsEventDefinitionId') AS DefId,
       JSON_VALUE(DataJson,'$.category') AS Category
FROM GenericDataStores
WHERE StoreName='operationsEvents'
  AND RecordId=
      'OPS-EVENT-MNT-SRESP-MNT-RESP-20260416115813-1-OED-MNT-START-413';""",
"category = 'MaintenanceExecution'.\nDefinition = OED-MNT-START.\nChain terminates at Equipment EQ-MIX-01.\nAll records have non-empty sourceTimeStamp.",
),
# ── TC-015 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-015",
"Verify that the latest sensor reading for every property of EQ-OVEN-01 is retrievable in a single query. Validates the multi-property latest-value pattern for equipment with 5 tracked properties.",
"Equipment → Equipment Property (latest record per property, all properties)",
"Equipment",
# Test Parameters
"Equipment ID: EQ-OVEN-01\n"
"Query type: latest record per each of 5 properties\n\n"
"Latest snapshot timestamp: 2026-05-05T13:48:00.000Z\n"
"Segment Response: SEG-RESP-PLANT01MUNICH-LINE-01-202605051151-RUN13-722",
# Expected Results
"Counts:\n  Equipment Property records returned: 5 (1 per property)\n\n"
"Latest values at 2026-05-05T13:48:00.000Z:\n"
"  EP-OVEN-TEMP  | OvenChamberTempC | 94.19 C\n"
"  EP-OVEN-HUM   | OvenHumidityPct  | 47.92 %\n"
"  EP-OVEN-PWR   | OvenPowerKW      | 26.62 kW\n"
"  EP-OVEN-CTRL  | ControlOutputPct | 44.08 %\n"
"  EP-OVEN-STATE | MachineState     | run EA",
"""-- Repeat for each property (replace EP-OVEN-TEMP as needed):
SELECT TOP 1 RecordId, CAST(DataJson AS NVARCHAR(MAX)) AS FullJson
FROM GenericDataStores
WHERE StoreName='equipmentPropertyTracking'
  AND RecordId LIKE
      'PROP-TRACK-EQ-ACT-PLANT01MUNICH-LINE-01-2026%EQ-OVEN-01%EP-OVEN-TEMP%'
ORDER BY RecordId DESC;
-- Replace EP-OVEN-TEMP with:
--   EP-OVEN-HUM, EP-OVEN-PWR, EP-OVEN-CTRL, EP-OVEN-STATE""",
"Exactly 5 records returned (one per property).\nAll belong to EQ-OVEN-01.\nValues match snapshot (94.19 / 47.92 / 26.62 / 44.08 / run).\nAll createdTimestamps = '2026-05-05 13:48:00'.",
),
# ── TC-016 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-016",
"Verify time-range filtering across all three production equipment items and their assigned properties simultaneously. Validates multi-equipment, multi-property bounded time queries (boundary-inclusive).",
"Equipment (×3) → Equipment Property (time-range filter, all properties)",
"Equipment IDs",
# Test Parameters
"Equipment IDs: EQ-MIX-01, EQ-OVEN-01, EQ-PACKAGE-01\n\n"
"Time window (inclusive):\n"
"  From: 2026-01-01T06:30:00.000Z\n"
"  To:   2026-01-01T08:30:00.000Z",
# Expected Results
"Counts per equipment:\n\n"
"EQ-MIX-01 — Total: 5,532\n"
"  EP-MIX-CUR:   1,442  (5 s interval)\n"
"  EP-MIX-VISC:    722  (10 s interval)\n"
"  EP-MIX-TORQ:  1,442  (5 s interval)\n"
"  EP-VIB:       1,442  (5 s interval)\n"
"  EP-MIX-STATE:   242  (state changes only)\n\n"
"EQ-OVEN-01 — Total: 6\n"
"  1 record per property at window boundary 08:30:00\n"
"  (Equipment starts at 08:30 UTC)\n\n"
"EQ-PACKAGE-01 — Total: 0\n"
"  (Line starts at 10:30 UTC — outside window)",
"""SELECT
    CASE
      WHEN RecordId LIKE '%EQ-MIX-01%'     THEN 'EQ-MIX-01'
      WHEN RecordId LIKE '%EQ-OVEN-01%'    THEN 'EQ-OVEN-01'
      WHEN RecordId LIKE '%EQ-PACKAGE-01%' THEN 'EQ-PACKAGE-01'
    END AS EquipmentId,
    COUNT(*) AS RecordCount
FROM GenericDataStores
WHERE StoreName='equipmentPropertyTracking'
  AND RecordId LIKE 'PROP-TRACK-EQ-ACT-PLANT01MUNICH-LINE-01-20260101%'
  AND (RecordId LIKE '%EQ-MIX-01%'
    OR RecordId LIKE '%EQ-OVEN-01%'
    OR RecordId LIKE '%EQ-PACKAGE-01%')
  AND RIGHT(RecordId,14) BETWEEN '20260101063000' AND '20260101083000'
GROUP BY
    CASE
      WHEN RecordId LIKE '%EQ-MIX-01%'     THEN 'EQ-MIX-01'
      WHEN RecordId LIKE '%EQ-OVEN-01%'    THEN 'EQ-OVEN-01'
      WHEN RecordId LIKE '%EQ-PACKAGE-01%' THEN 'EQ-PACKAGE-01'
    END;""",
"EQ-MIX-01: 5,532 (EP-MIX-STATE fewer due to state-change-only sampling).\nEQ-OVEN-01: 6 (1 per property at window upper bound).\nEQ-PACKAGE-01: 0.\nNo records outside the window range.",
),
# ── TC-017 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-017",
"Verify that parent equipment resolves to its child equipment, and each child's latest property value is accessible. Also verifies that a child with no property assignments returns an empty set.",
"Equipment → Child Equipment (ParentEquipmentId) → Child Equipment Property (latest)",
"Parent Equipment IDs",
# Test Parameters
"Parent Equipment IDs:\n  EQ-MIX-01, EQ-OVEN-01, EQ-PACKAGE-01\n\n"
"Query type: latest child property value\n\n"
"Snapshot timestamps:\n"
"  EQ-MIX-MOTOR-01: 2026-05-05T09:51:00.000Z\n"
"  EQ-OVEN-ZONE-01: 2026-05-05T13:48:00.000Z",
# Expected Results
"Counts:\n  Parent Equipment: 3\n  Child Equipment: 3\n  Children with property assignments: 2\n\n"
"Child latest values:\n"
"  EQ-MIX-MOTOR-01 → EP-MIX-MOTORSPEED\n"
"    Property: MixerMotorSpeed\n"
"    Value: 8.31 RPM\n"
"    Timestamp: 2026-05-05T09:51:00.000Z\n\n"
"  EQ-OVEN-ZONE-01 → EP-OVEN-ZONE-HEAT\n"
"    Property: Oven Zone Heat\n"
"    Value: 99.97 Celsius\n"
"    Timestamp: 2026-05-05T13:48:00.000Z\n\n"
"  EQ-PACK-LABELER-01 → no property assignments (returns empty)",
"""-- Child equipment lookup:
SELECT Id, Name, ParentEquipmentId
FROM Equipments
WHERE ParentEquipmentId IN ('EQ-MIX-01','EQ-OVEN-01','EQ-PACKAGE-01');
-- Labeler property assignment count (expect 0):
SELECT COUNT(*) AS AssignmentCount
FROM EquipmentPropertyAssignments
WHERE EquipmentId='EQ-PACK-LABELER-01';
-- Latest Motor Speed:
SELECT TOP 1 RecordId, CAST(DataJson AS NVARCHAR(MAX)) AS FullJson
FROM GenericDataStores WHERE StoreName='equipmentPropertyTracking'
  AND RecordId LIKE
      'PROP-TRACK-EQ-ACT-PLANT01MUNICH-LINE-01-202605%EQ-MIX-MOTOR-01%EP-MIX-MOTORSPEED%'
ORDER BY RecordId DESC;
-- Latest Zone Heat:
SELECT TOP 1 RecordId, CAST(DataJson AS NVARCHAR(MAX)) AS FullJson
FROM GenericDataStores WHERE StoreName='equipmentPropertyTracking'
  AND RecordId LIKE
      'PROP-TRACK-EQ-ACT-PLANT01MUNICH-LINE-01-202605%EQ-OVEN-ZONE-01%EP-OVEN-ZONE-HEAT%'
ORDER BY RecordId DESC;""",
"Each parent resolves to exactly 1 child.\nEQ-MIX-MOTOR-01: EP-MIX-MOTORSPEED = 8.31 RPM @ 2026-05-05T09:51:00Z.\nEQ-OVEN-ZONE-01: EP-OVEN-ZONE-HEAT = 99.97 Celsius @ 2026-05-05T13:48:00Z.\nEQ-PACK-LABELER-01: 0 assignments (returns empty).",
),
# ── TC-018 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-018",
"Verify that the direct children of an Area-level Hierarchy Scope node are resolvable. Validates the first hop of ISA-95 hierarchy graph navigation.",
"Hierarchy Scope (Area) → Direct Child Hierarchy Scopes",
"Hierarchy Scope",
# Test Parameters
"Hierarchy Scope ID:\n  HS-PLANT01MUNICH-BAKING-MUNICH-Area\n\n"
"EquipmentLevel: Area\n"
"EquipmentID: BAKING-MUNICH",
# Expected Results
"Counts:\n  HierarchyScopeParentChilds entries: 1\n  Child Hierarchy Scopes: 1\n\n"
"Child entry:\n"
"  HSPC-0003:\n"
"    parent = HS-PLANT01MUNICH-BAKING-MUNICH-Area\n"
"    child  = HS-PLANT01MUNICH-WCenter1-WorkCenter\n"
"    child level = Work Center",
"""SELECT hspc.Id, hspc.ChildEquipmentID, hspc.ChildEquipmentLevel,
       hs.EquipmentID, hs.EquipmentLevel
FROM HierarchyScopeParentChilds hspc
JOIN HierarchyScopes hs ON hs.Id = hspc.ChildEquipmentID
WHERE hspc.ParentEquipmentID = 'HS-PLANT01MUNICH-BAKING-MUNICH-Area';""",
"Exactly 1 HSPC entry with this parent.\nChild node EquipmentLevel = Work Center.\nAll records have non-empty sourceTimeStamp.",
),
# ── TC-019 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-019",
"Verify that all equipment (Work Cells) belonging to a Production Line are reachable via a 2-hop hierarchy traversal through the Production Unit. Validates equipment discovery from line context.",
"Hierarchy Scope (Production Line) → Production Unit → Equipment Work Cells",
"Hierarchy Scope",
# Test Parameters
"Hierarchy Scope ID:\n  HS-PLANT01MUNICH-LINE-01-ProductionLine\n\n"
"EquipmentID: LINE-01\n"
"EquipmentLevel: Production Line\n\n"
"Traversal depth: 2 hops",
# Expected Results
"Counts:\n  Intermediate Production Unit hops: 1 (HSPC-0008)\n  Equipment Work Cells: 3\n\n"
"Work Cells resolved:\n"
"  HS-PLANT01MUNICH-EQ-MIX-01-WorkCell    → EQ-MIX-01\n"
"  HS-PLANT01MUNICH-EQ-OVEN-01-WorkCell   → EQ-OVEN-01\n"
"  HS-PLANT01MUNICH-EQ-PACKAGE-01-WorkCell→ EQ-PACKAGE-01\n\n"
"HSPC entries used: HSPC-0008, HSPC-0009, HSPC-0012, HSPC-0015",
"""SELECT hs_wc.Id, hs_wc.EquipmentID, e.Name
FROM HierarchyScopeParentChilds hspc_line
JOIN HierarchyScopeParentChilds hspc_punit
     ON hspc_punit.ParentEquipmentID = hspc_line.ChildEquipmentID
     AND hspc_punit.ChildEquipmentLevel = 'Work Cell'
JOIN HierarchyScopes hs_wc ON hs_wc.Id = hspc_punit.ChildEquipmentID
JOIN Equipments e ON e.Id = hs_wc.EquipmentID
WHERE hspc_line.ParentEquipmentID =
      'HS-PLANT01MUNICH-LINE-01-ProductionLine'
ORDER BY hs_wc.EquipmentID;""",
"Production Unit hop = HSPC-0008.\n3 Work Cell hops = HSPC-0009, HSPC-0012, HSPC-0015.\nEach Work Cell EquipmentID resolves to an existing Equipment record.",
),
# ── TC-020 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-020",
"Verify that all Production Lines belonging to a Site can be found by recursively traversing the hierarchy graph. Validates deep recursive traversal (6 hops Site → Production Line).",
"Hierarchy Scope (Site) → Descendant Production Lines (recursive, 6 hops)",
"Hierarchy Scope",
# Test Parameters
"Hierarchy Scope ID:\n  HS-PLANT01MUNICH-Site\n\n"
"EquipmentID: PLANT01MUNICH\n"
"EquipmentLevel: Site\n\n"
"Target level: Production Line\n"
"Traversal depth: 6 hops",
# Expected Results
"Counts:\n  Production Line descendants: 1\n\n"
"Production Line: LINE-01\n\n"
"Traversal path:\n"
"  Depth 1 – HSPC-0002: Site → Area\n"
"  Depth 2 – HSPC-0003: Area → Work Center\n"
"  Depth 3 – HSPC-0004: Work Center → Work Unit\n"
"  Depth 4 – HSPC-0005: Work Unit → Process Cell\n"
"  Depth 5 – HSPC-0006: Process Cell → Unit\n"
"  Depth 6 – HSPC-0007: Unit → HS-PLANT01MUNICH-LINE-01-ProductionLine",
"""WITH hierarchy AS (
    SELECT ChildEquipmentID AS NodeId,
           ChildEquipmentLevel AS NodeLevel,
           1 AS Depth
    FROM HierarchyScopeParentChilds
    WHERE ParentEquipmentID = 'HS-PLANT01MUNICH-Site'
    UNION ALL
    SELECT c.ChildEquipmentID, c.ChildEquipmentLevel, h.Depth+1
    FROM HierarchyScopeParentChilds c
    INNER JOIN hierarchy h ON c.ParentEquipmentID = h.NodeId
)
SELECT h.NodeId AS ProductionLineHsId,
       hs.EquipmentID AS LineId,
       h.Depth
FROM hierarchy h
JOIN HierarchyScopes hs ON hs.Id = h.NodeId
WHERE h.NodeLevel = 'Production Line';""",
"Exactly 1 Production Line descendant.\nLineId = LINE-01.\nTraversal depth = 6.\nAll 6 HSPC entries (HSPC-0002 through HSPC-0007) used.",
),
# ── TC-021 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-021",
"Verify that an Equipment record resolves to its Equipment Class and that all 7 class-level properties are accessible. Validates Equipment-to-Class property navigation for MIXER equipment.",
"Equipment → Equipment Class → Equipment Class Properties (MIXER)",
"Equipment",
# Test Parameters
"Equipment ID: EQ-MIX-01\n"
"Equipment Class: MIXER",
# Expected Results
"Counts:\n  Equipment Class Properties: 7\n\n"
"Class Properties:\n"
"  ECP-MIX-CTRL  – ControlOutputPct\n"
"  ECP-MIX-CUR   – MixerMotorCurrentA\n"
"  ECP-MIX-STATE – MachineState\n"
"  ECP-MIX-TORQ  – MixerTorqueNm\n"
"  ECP-MIX-VISC  – DoughViscosityIndex\n"
"  ECP-PRESSURE  – OperatingPressure\n"
"  ECP-VIB       – EquipmentVibrationRMS\n\n"
"Sample ECPAs: ECP-MIX-CUR_EP-MIX-CUR, ECP-VIB_EP-VIB",
"""SELECT e.Id, e.ClassId,
       ecp.Id AS ClassPropertyId, ecp.Name
FROM Equipments e
JOIN EquipmentClassProperties ecp
     ON ecp.EquipmentClassId = e.ClassId
WHERE e.Id = 'EQ-MIX-01'
ORDER BY ecp.Id;
SELECT ecpa.Id,
       ecpa.EquipmentClassPropertyId,
       ecpa.EquipmentPropertyId
FROM EquipmentClassPropertyAssignments ecpa
WHERE ecpa.EquipmentClassPropertyId IN (
  SELECT Id FROM EquipmentClassProperties
  WHERE EquipmentClassId='MIXER');""",
"ClassId = 'MIXER'.\nExactly 7 EquipmentClassProperties for MIXER.\nECPAs use pattern {ECP-ID}_{EP-ID}.\nAll records have non-empty UpdatedAt.",
),
# ── TC-022 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-022",
"Verify that an Oven Equipment record resolves to its Equipment Class and that all 6 class-level properties are accessible. Validates Equipment-to-Class property navigation for OVEN equipment.",
"Equipment → Equipment Class → Equipment Class Properties (OVEN)",
"Equipment",
# Test Parameters
"Equipment ID: EQ-OVEN-01\n"
"Equipment Class: OVEN",
# Expected Results
"Counts:\n  Equipment Class Properties: 6\n\n"
"Class Properties:\n"
"  ECP-OVEN-CTRL  – ControlOutputPct\n"
"  ECP-OVEN-HUM   – OvenHumidityPct\n"
"  ECP-OVEN-PWR   – OvenPowerKW\n"
"  ECP-OVEN-STATE – MachineState\n"
"  ECP-OVEN-TEMP  – OvenChamberTempC\n"
"  ECP-TEMPERATURE– OperatingTemperature\n\n"
"Sample ECPAs: ECP-OVEN-TEMP_EP-OVEN-TEMP, ECP-OVEN-STATE_EP-OVEN-STATE",
"""SELECT e.Id, e.ClassId,
       ecp.Id AS ClassPropertyId, ecp.Name
FROM Equipments e
JOIN EquipmentClassProperties ecp
     ON ecp.EquipmentClassId = e.ClassId
WHERE e.Id = 'EQ-OVEN-01'
ORDER BY ecp.Id;""",
"ClassId = 'OVEN'.\nExactly 6 EquipmentClassProperties for OVEN.\nAll records have non-empty UpdatedAt.",
),
# ── TC-023 ──────────────────────────────────────────────────────────────────
(
"TC-FUNC-REL-023",
"Verify that a Packaging Equipment record resolves to its Equipment Class and that both class-level properties are accessible. Validates Equipment-to-Class property navigation for PACK (minimal property set).",
"Equipment → Equipment Class → Equipment Class Properties (PACK)",
"Equipment",
# Test Parameters
"Equipment ID: EQ-PACKAGE-01\n"
"Equipment Class: PACK",
# Expected Results
"Counts:\n  Equipment Class Properties: 2\n\n"
"Class Properties:\n"
"  ECP-PACK-CTRL  – ControlOutputPct\n"
"  ECP-PACK-STATE – MachineState\n\n"
"Sample ECPA: ECP-PACK-STATE_EP-PACK-STATE",
"""SELECT e.Id, e.ClassId,
       ecp.Id AS ClassPropertyId, ecp.Name
FROM Equipments e
JOIN EquipmentClassProperties ecp
     ON ecp.EquipmentClassId = e.ClassId
WHERE e.Id = 'EQ-PACKAGE-01'
ORDER BY ecp.Id;""",
"ClassId = 'PACK'.\nExactly 2 EquipmentClassProperties for PACK.\nAll records have non-empty UpdatedAt.",
),
]

# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

def build_workbook(output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.fill      = HEADER_FILL
        c.font      = WHITE
        c.alignment = MID
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Second header row: sub-labels
    sub_labels = [
        "", "", "", "",
        "Anchor Entity ID  |  Timestamp / Time Window",
        "Entity Counts  |  Key IDs  |  Latest Values + UoM",
        "",
        "",
    ]
    for col_idx, label in enumerate(sub_labels, start=1):
        c = ws.cell(row=2, column=col_idx, value=label)
        c.fill      = PatternFill("solid", fgColor="2E75B6")
        c.font      = Font(name="Calibri", size=9, color="FFFFFF", italic=True)
        c.alignment = MID
        c.border    = BORDER

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, tc in enumerate(TC, start=3):
        is_odd = (row_idx % 2 == 1)
        base_fill   = ODD_FILL    if is_odd else EVEN_FILL
        param_fill  = PARAM_FILL_ODD  if is_odd else PARAM_FILL_EVEN
        exp_fill    = EXP_FILL_ODD    if is_odd else EXP_FILL_EVEN

        (tc_id, func_desc, rel_chain, anchor_type,
         test_params, expected_results,
         sql, criteria) = tc

        row_data = [
            (COL_TC,       tc_id,            base_fill,  TCID,   WRAP),
            (COL_DESC,     func_desc,         base_fill,  NORMAL, WRAP),
            (COL_CHAIN,    rel_chain,         base_fill,  NORMAL, WRAP),
            (COL_TYPE,     anchor_type,       base_fill,  NORMAL, WRAP),
            (COL_PARAMS,   test_params,       param_fill, NORMAL, WRAP),
            (COL_EXPECTED, expected_results,  exp_fill,   NORMAL, WRAP),
            (COL_SQL,      sql,               SQL_FILL,   MONO,   WRAP),
            (COL_CRITERIA, criteria,          base_fill,  NORMAL, WRAP),
        ]

        for col_idx, value, fill, font, align in row_data:
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.fill      = fill
            c.font      = font
            c.alignment = align
            c.border    = BORDER

        ws.row_dimensions[row_idx].height = 110

    # ── Legend tab ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Legend")
    legend = [
        ("Column",               "Description"),
        ("TC ID",                "Unique test case identifier (TC-FUNC-REL-NNN)"),
        ("Functional Description",
         "Plain-English description of what the test validates from a business/functional perspective"),
        ("Relationship Chain",
         "Technical ISA-95 entity chain traversed in the test (A → B → C)"),
        ("Anchor Entity Type",
         "The ISA-95 entity type used as the starting filter point"),
        ("Test Parameters",
         "Anchor Entity ID(s) that drive the test, plus relevant timestamps or time windows. "
         "Green background — these are the INPUTS to the test."),
        ("Expected Results",
         "Expected entity counts at each chain hop, key IDs/names, and latest sensor values with UoM. "
         "All values sourced from ISA95Migrations_SQL2022 snapshot (2026-04-17, updated 2026-05-07). "
         "Orange background — these are the EXPECTED OUTPUTS of the test."),
        ("Validation SQL",
         "SQL query against ISA95Migrations_SQL2022 to produce the expected output. "
         "Store name: equipmentPropertyTracking (camelCase). "
         "Timestamp field inside DataJson: createdTimestamp."),
        ("Validation Criteria",
         "Specific assertions that must hold for the test to PASS."),
    ]

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 100
    for r, (k, v) in enumerate(legend, start=1):
        ck = ws2.cell(row=r, column=1, value=k)
        cv = ws2.cell(row=r, column=2, value=v)
        ck.font      = BOLD
        cv.font      = NORMAL
        ck.alignment = WRAP
        cv.alignment = WRAP
        ws2.row_dimensions[r].height = 36
        if r == 1:
            ck.fill = HEADER_FILL; ck.font = WHITE
            cv.fill = HEADER_FILL; cv.font = WHITE
        elif k == "Test Parameters":
            ck.fill = PARAM_FILL_ODD; cv.fill = PARAM_FILL_ODD
        elif k == "Expected Results":
            ck.fill = EXP_FILL_ODD; cv.fill = EXP_FILL_ODD

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    out = os.path.join(os.path.dirname(__file__), "ISA95_Test_Cases.xlsx")
    build_workbook(out)
